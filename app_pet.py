# -*- coding: utf-8 -*-
"""
Created on Sat Sep 21 19:02:42 2019

@author: EJ
"""

# -*- coding: utf-8 -*-

import requests
from flask import Flask, request, Response
from pet_pharmacy import pet_pharm_api
from openpyxl import load_workbook

EXCEL_FILE_NAME = 'Database.xlsx'
db = load_workbook(filename=EXCEL_FILE_NAME)
tuto_db = db['fv']

API_KEY = '916473331:AAHDtuAaRcGM8FlPRFUYDGRzY6mgyDWacuo' #gonnabe
#API_KEY = '764462849:AAER9m2z9X4jRkQ4SYycPUo91o16fXxIzhk' #iluvpet
app = Flask(__name__)

    
def write_title_list(userTitle):
    tuto_db['A3'].value = userTitle
    db.save(EXCEL_FILE_NAME)
    
def write_detail_list(userLocList):
    tuto_db['A4'].value = userLocList
    db.save(EXCEL_FILE_NAME)
 
def read_with_index(loc):
    read_result = tuto_db[loc].value
    return read_result

def read_with_index_faq():
    load_wb = load_workbook("Database.xlsx", data_only=True)
    fv_sheet = load_wb['fv']  
    return fv_sheet

def faq_answer(get_cells):
    answer=''
    for row in get_cells:
        for cell in row:
            answer=answer+cell.value+'\n'           
    return answer
    
def read_with_sm_pet_hospital(user_pick_name='차오름동물병원'):
    sm_pet=load_workbook(filename="small_pet_database.xlsx", data_only=True)
    sm_db=sm_pet['Sheet1']
    all_values = []
    for row in sm_db.rows:
        row_value = []
        for cell in row:
            row_value.append(cell.value)
        all_values.append(row_value)
    
    #서울특별시내 소동물병원 목록.
    sm_title_list=''
    for item in all_values[1:]:
      i=1
      sm_title_list=sm_title_list+item[i-1] +' ('+item[i]+')'+'\n'
      i=i+1
      
    #선택한 병원 디테일정보
    detail_info=''

    for item in all_values:
        if user_pick_name in item[0]: 
            detail_info=detail_info+str(item)
            detail_info_list=detail_info.split(',')          
            name=all_values[0][0]+' - '+detail_info_list[0][1:]+'\n'
            gu=all_values[0][1]+' - '+detail_info_list[1]+'\n'
            time=all_values[0][2]+' - '+detail_info_list[2]+'\n'
            time_weekend=all_values[0][3]+'-'+detail_info_list[3]+'\n'
            tel=all_values[0][4]+' - '+detail_info_list[4]+'\n'
            addr=all_values[0][5]+' - '+detail_info_list[5]+'\n'
            ref=all_values[0][6]+' - '+detail_info_list[6][:-2]
            result = name+gu+time+time_weekend+tel+addr+ref
            little_list=sm_title_list[:121]
    return result,sm_title_list,little_list 

def parse_message(data):
    '''응답data 로부터 chat_id 와 text, user_name을 추출.'''
    chat_id = None
    msg = None
    user_name = None
    inline_data = None    
    
    if 'callback_query' in data:
        data=data['callback_query']
        inline_data = data['data']
        
    chat_id = data['message']['chat']['id']
    msg = data['message']['text']
    user_name = data['message']['chat']['first_name'] + data['message']['chat']['last_name']
    print('parse_message안에서 inline data==',inline_data)
    return chat_id, msg, user_name, inline_data    #https://core.telegram.org/bots/api#keyboardbutton

def send_message_map_url(chat_id,text):
    url = 'https://api.telegram.org/bot{token}/sendMessage'.format(token=API_KEY)  
    google_map_url='https://www.google.com/maps/search/?api=1&query='
    google_map_user_url=(google_map_url+text).replace(" ","")
    params = {'chat_id':chat_id, 'text':google_map_user_url}
    requests.post(url, json=params)

def send_message_inlinekeyboard(chat_id, text):
    url = 'https://api.telegram.org/bot{token}/sendMessage'.format(token=API_KEY)  
    InlineKeyboard = {'inline_keyboard' : [[{'text':'전체보기', 'callback_data':'ALLVIEW'}]]}
    params = {'chat_id':chat_id, 'text' : '전체 목록을 보시려면 하단 버튼을 클릭하세요' , 'reply_markup':InlineKeyboard}
    requests.post(url,json=params)

def send_message_inlinekeyboard_pharmacy(chat_id, text):
    url = 'https://api.telegram.org/bot{token}/sendMessage'.format(token=API_KEY)  
    InlineKeyboard = {'inline_keyboard' : [[{'text':'전체보기', 'callback_data':'ALLVIEW_PHARMACY'}] ]}
    params = {'chat_id':chat_id, 'text' : '전체 목록을 보시려면 하단 버튼을 클릭하세요' , 'reply_markup':InlineKeyboard}
    requests.post(url,json=params)
    
def send_message(chat_id, text='bla-bla-bla', user_name='noone', inline_data='hi'):
    url = 'https://api.telegram.org/bot{token}/sendMessage'.format(token=API_KEY)   #sendMessage
    keyboard = {                                      
            'keyboard':[[{
                    'text': '동물약국 찾기'
                        },
                    {'text': '소동물병원 찾기'
                        }],[
                                
                    {'text': '과태료 정보'
                        },{'text': 'FAQ'
                        }]
                    ],
            'one_time_keyboard' : True
            }
    keyboard_FAQ = {                                      
            'keyboard':[
                    [{
                    'text': '동물등록방법'
                        }],
                    [{
                    'text': '반려동물을 잃어버렸어요.'
                        }],
                    [{
                    'text': '길잃은 동물을 봤어요.'
                        }],
                [{
                    'text': '로드킬 신고는 어디에 하나요?'
                        }],
                [{
                    'text': '동물사체처리는 어떻게 하나요?'
                        }],
                [{
                    'text': '처음화면 가기'
                        }]
                    ],
            'one_time_keyboard' : True
            }
    if  text[:5] == "지역검색!":
        user_loc=text[5:] # ex)노원구 
        print('------------사용자가 지역검색-------:',user_loc)
        userTitle,userLocList =  pet_pharm_api(user_loc) #함수호출
        write_title_list(userTitle)#엑셀에 작성하는 코드 A3에 타이틀리스트 저장  
        write_detail_list(userLocList)#엑셀에 작성하는 코드 A4에 디테일리스트 저장  
        read_title_A3=read_with_index('A3')      
        short_result=read_title_A3[:219]
        result=user_loc+'에 있는 동물약국 목록 입니다. 상세정보를 원하는 약국이름을 입력해주세요. 입력방식=[더보기!+약국이름] \n 예)더보기!봄약국' +'\n'+'\n'+short_result +'.. ...등등'     
        params = {'chat_id':chat_id, 'text': result}
        requests.post(url, json=params)
        send_message_inlinekeyboard_pharmacy(chat_id, text)
    elif text=='FAQ':
         params = {'chat_id':chat_id, 'text': '동물관련 FAQ입니다. 궁금하신 사항을 선택해주세요.', 'reply_markup' : keyboard_FAQ}
         requests.post(url, json=params)
    elif text== '동물등록방법':
         fv_sheet=read_with_index_faq()         
         get_cells = fv_sheet['A50':'A56']
         final_answer=faq_answer(get_cells)                    
         params = {'chat_id':chat_id, 'text':final_answer, 'reply_markup' : keyboard_FAQ}
         requests.post(url, json=params)
    elif text=='반려동물을 잃어버렸어요.':
         fv_sheet=read_with_index_faq()         
         get_cells = fv_sheet['B50':'B52']
         final_answer=faq_answer(get_cells)                    
         params = {'chat_id':chat_id, 'text':final_answer, 'reply_markup' : keyboard_FAQ}
         requests.post(url, json=params)
    elif text=='길잃은 동물을 봤어요.':
         fv_sheet=read_with_index_faq()         
         get_cells = fv_sheet['C50':'C51']
         final_answer=faq_answer(get_cells)                    
         params = {'chat_id':chat_id, 'text':final_answer, 'reply_markup' : keyboard_FAQ}
         requests.post(url, json=params)    
    elif text=='로드킬 신고는 어디에 하나요?':
         fv_sheet=read_with_index_faq()         
         get_cells = fv_sheet['D50':'D51']
         final_answer=faq_answer(get_cells)                    
         params = {'chat_id':chat_id, 'text':final_answer, 'reply_markup' : keyboard_FAQ}
         requests.post(url, json=params)   
    elif text=='동물사체처리는 어떻게 하나요?':
         fv_sheet=read_with_index_faq()         
         get_cells = fv_sheet['E50':'E51']
         final_answer=faq_answer(get_cells)                    
         params = {'chat_id':chat_id, 'text':final_answer, 'reply_markup' : keyboard_FAQ}
         requests.post(url, json=params)        
    elif text == '소동물병원 찾기':
         detail_info,sm_hospital_total_list,little_list = read_with_sm_pet_hospital()
         params = {'chat_id':chat_id, 'text': '서울특별시 소동물병원 리스트 입니다. 상세정보를 원하시는 병원명을 입력하세요. \n 예)힐스타동물병원  \n\n'+little_list} 
         requests.post(url, json=params)
         send_message_inlinekeyboard(chat_id, text)
         
    elif inline_data=='ALLVIEW':
         detail_info,sm_hospital_total_list,little_list = read_with_sm_pet_hospital()
         params = {'chat_id':chat_id, 'text': '서울특별시 소동물병원 리스트 입니다. 상세정보를 원하시는 병원명을 입력하세요.\n 예)힐스타동물병원   \n\n'+sm_hospital_total_list+'\n \n처음으로 돌아가시려면 "처음"을 입력해주세요.'}  
         requests.post(url, json=params)
         
    elif inline_data=='ALLVIEW_PHARMACY':
         read_title_A3=read_with_index('A3')
         params = {'chat_id':chat_id, 'text': '동물약 취급 약국 전체 리스트 입니다. 상세정보를 원하시는 약국이름을 입력하세요. 입력방식=[더보기!+약국이름] \n 예)더보기!봄약국   \n\n'+read_title_A3}  
         requests.post(url, json=params)
          
    elif text == '과태료 정보':     
         safe_rule = tuto_db['A20':'A42']
         safe_fine=''
         for row in safe_rule:
                for cell in row:
                    safe_fine=safe_fine+str(cell.value)+'\n'
         safe_fine_str=str(safe_fine)           
         params = {'chat_id':chat_id, 'text': safe_fine_str+'\n \n처음으로 돌아가시려면 "처음"을 입력해주세요.'}
         requests.post(url, json=params)     
         
    elif (text[-2:]=='병원') | (text[-2:]=='센터'):
         user_hopital_name=text #병원이름 들어감
         detail_info,sm_hospital_total_list,little_list=read_with_sm_pet_hospital(user_hopital_name)
         params = {'chat_id':chat_id, 'text': detail_info+'\n \n처음화면으로 돌아가시려면 "처음"을 입력해주세요.'}
         requests.post(url, json=params)
         
    elif text[:4]=='더보기!':
        num=text[4:]# 예)병원이름..
       
        #엑셀에서 엔터로 구분해서 읽으들어온다.str->list
        db_a4_str=read_with_index('A4') #A4셀에 저장해두었던 상세정보 읽어온다.
        db_a4_list=db_a4_str.split('\n')#엔터로 구분해서 list로 변환.
      
        #user가 선택한 번호를 db_a4_list에서 찾아서 디테일정보를 가져온다.     
        for item in db_a4_list:
            if num in item:
                detail=item
                    
        params = {'chat_id':chat_id, 'text': detail +'\n \n처음화면으로 돌아가시려면 "처음"을 입력해주세요.'}
        requests.post(url, json=params)
        
        #구글 지도 유알엘 출력해주는 센더함수 
        sep_addr=detail.split('주소:')
        sep_slash=sep_addr[1].split('/')
        final_addr=sep_slash[0]
        send_message_map_url(chat_id,text=final_addr) 
        
    elif text=="동물약국 찾기":
        params = {'chat_id':chat_id, 'text': '검색을 원하는 지역을 말씀해주세요! \n 입력방식=[지역검색!+지역구] \n 예)지역검색!노원구'}
        requests.post(url, json=params)
        
    elif (text=='처음')|(text=='처음화면 가기'):      
        params = {'chat_id':chat_id, 'text': '안녕하세요. '+user_name+'님! '+'\n펫을 위한 모든 정보! 🐹️아이러브펫🐹️입니다. 이용방법은 하단 설명을 참고해주세요.\n\n * 동물약국 검색은 [동물약국찾기] 버튼을 클릭해주세요. \n* 햄스터,기니피그,토끼,친칠라 등 을 위한 병원은 [소동물찾기] 버튼을 클릭해주세요.  \n* 동물등록방법, 반려동물 분실 및 신고, 로드킬 신고 방법 등은  [FAQ] 버튼을 클릭해주세요. \n* 과태료 관련 정보를 확인 하고 싶으시면 [과태료 정보]를 클릭해주세요. \n* 처음화면으로 돌아가고 싶으면 언제든 [처음]을 입력해주세요. ', 'reply_markup' : keyboard}
        requests.post(url, json=params)
   
    else:
        params = {'chat_id':chat_id, 'text': '안녕하세요. '+user_name+'님! '+'\n펫을 위한 모든 정보! 🐹️아이러브펫🐹️입니다. 이용방법은 하단 설명을 참고해주세요.\n\n * 동물약국 검색은 [동물약국찾기] 버튼을 클릭해주세요. \n* 햄스터,기니피그,토끼,친칠라 등 을 위한 병원은 [소동물찾기] 버튼을 클릭해주세요.  \n* 동물등록방법, 반려동물 분실 및 신고, 로드킬 신고 방법 등은  [FAQ] 버튼을 클릭해주세요. \n* 과태료 관련 정보를 확인 하고 싶으시면 [과태료 정보]를 클릭해주세요. \n* 처음화면으로 돌아가고 싶으면 언제든 [처음]을 입력해주세요. ', 'reply_markup' : keyboard}
        requests.post(url, json=params)

    
    return 0


    
# 경로 설정, URL 설정
@app.route('/', methods=['POST', 'GET'])
def index():
    if request.method == 'POST':
        message = request.get_json()           
        chat_id, msg, user_name, inline_data = parse_message(message)
        send_message(chat_id, msg, user_name, inline_data)                       
        return Response('ok', status=200)
    else:
        return 'Hello World!'


    

if __name__ == '__main__':
    app.run(port = 5000)